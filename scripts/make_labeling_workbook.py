from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from common import COMMITMENT_LABELS, EMOTION_LABELS, RECOMMENDATION_LABELS


HEADERS = [
    "review_uid",
    "rating",
    "word_count",
    "length_bucket",
    "review_text",
    "coder_1_emotions",
    "coder_1_commitment",
    "coder_1_recommendation",
    "coder_1_quality_issue",
    "coder_1_notes",
    "coder_2_emotions",
    "coder_2_commitment",
    "coder_2_recommendation",
    "coder_2_quality_issue",
    "coder_2_notes",
    "coder_3_emotions",
    "coder_3_commitment",
    "coder_3_recommendation",
    "coder_3_quality_issue",
    "coder_3_notes",
]


def add_validations(ws, start_row: int, end_row: int) -> None:
    commitment = DataValidation(type="list", formula1=f'"{",".join(COMMITMENT_LABELS)}"', allow_blank=True)
    recommendation = DataValidation(type="list", formula1=f'"{",".join(RECOMMENDATION_LABELS)}"', allow_blank=True)
    quality = DataValidation(type="list", formula1='"false,true"', allow_blank=True)
    ws.add_data_validation(commitment)
    ws.add_data_validation(recommendation)
    ws.add_data_validation(quality)
    for coder_offset in [6, 11, 16]:
        commitment.add(f"{ws.cell(row=start_row, column=coder_offset + 1).coordinate}:{ws.cell(row=end_row, column=coder_offset + 1).coordinate}")
        recommendation.add(f"{ws.cell(row=start_row, column=coder_offset + 2).coordinate}:{ws.cell(row=end_row, column=coder_offset + 2).coordinate}")
        quality.add(f"{ws.cell(row=start_row, column=coder_offset + 3).coordinate}:{ws.cell(row=end_row, column=coder_offset + 3).coordinate}")


def format_sheet(ws, n_rows: int) -> None:
    header_fill = PatternFill("solid", fgColor="E9EEF5")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="172033")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "F2"
    widths = {
        "A": 18,
        "B": 8,
        "C": 10,
        "D": 13,
        "E": 90,
        "F": 28,
        "G": 16,
        "H": 20,
        "I": 16,
        "J": 28,
        "K": 28,
        "L": 16,
        "M": 20,
        "N": 16,
        "O": 28,
        "P": 28,
        "Q": 16,
        "R": 20,
        "S": 16,
        "T": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row[5:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = f"A1:T{n_rows + 1}"
    emotion_help = "Enter one or more emotion labels separated by commas. Allowed: " + ", ".join(EMOTION_LABELS) + ". Use none only by itself."
    for col in ["F", "K", "P"]:
        ws[f"{col}1"].comment = Comment(emotion_help, "Codex")


def write_label_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    ws.append(HEADERS)
    for _, row in df.iterrows():
        ws.append([
            row.get("review_uid", ""),
            row.get("rating", ""),
            row.get("word_count", ""),
            row.get("length_bucket", ""),
            row.get("review_text", ""),
            "", "", "", "", "",
            "", "", "", "", "",
            "", "", "", "", "",
        ])
    format_sheet(ws, len(df))
    add_validations(ws, 2, len(df) + 1)


def build(input_path: Path, output_path: Path, pilot_size: int) -> None:
    df = pd.read_csv(input_path)
    df = df[["review_uid", "rating", "word_count", "length_bucket", "review_text"]].copy()
    pilot = df.head(pilot_size).copy()
    holdout = df.copy()

    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    lines = [
        ["Goal", "Label each entire review for emotions, commitment, and recommendation."],
        ["Emotions", "Multi-label. Type comma-separated labels. Use none only when no emotion is present."],
        ["Allowed emotions", ", ".join(EMOTION_LABELS)],
        ["Commitment", "Dropdown: low, medium, high. Use medium if continuation is unclear or not mentioned."],
        ["Recommendation", "Dropdown: would_not, neutral, would_recommend. Use neutral if unclear, conditional, or not mentioned."],
        ["Quality issue", "Use true only for empty, spam, unintelligible, or non-review text."],
        ["Pilot", "Coders should label the Pilot sheet first, discuss disagreements, then label the Holdout sheet independently."],
    ]
    for row in lines:
        instructions.append(row)
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 110
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    instructions["A1"].font = Font(bold=True)

    write_label_sheet(wb, "Pilot", pilot)
    write_label_sheet(wb, "Holdout", holdout)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create an Excel labeling workbook with dropdowns and coder columns.")
    parser.add_argument("--input", default=Path("data/processed/holdout_locked.csv"), type=Path)
    parser.add_argument("--output", default=Path("data/labels/holdout_labeling_workbook.xlsx"), type=Path)
    parser.add_argument("--pilot-size", default=60, type=int)
    args = parser.parse_args()
    build(args.input, args.output, args.pilot_size)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()

